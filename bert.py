import pickle
from bert_serving.client import BertClient
bc = BertClient()
doc_vecs = bc.encode(['小黄车哈哈哈中文', 'then do it right', 'then do it better'])
print(doc_vecs)
print(doc_vecs.shape)

import openpyxl
import torch
import numpy as np
data = openpyxl.load_workbook("bertsentence.xlsx")
sheetnames = data.get_sheet_names()
table = data.get_sheet_by_name(sheetnames[0])
nrows = table.max_row  # 获得行数

begin = 2
end = 22349
list = []
while(begin<=end):
    sentence = str(table.cell(begin, 2).value)
    list.append(sentence)
    begin = begin + 1

begin = 2
end = 22349
doc_vec = bc.encode(list)

plist = {}
for i in doc_vec:
    print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")
    print(i[0])
    variablex = torch.tensor(i[0], dtype=torch.double)
    key = str(table.cell(begin, 1).value)
    plist[key] = variablex
    begin = begin + 1

data.save("bertsentence.xlsx")

with open("douyin_pbert.pkl", 'wb') as fo:  # 将数据写入pkl文件
    pickle.dump(plist, fo)


begin = 2
end = 22349
list = []
while(begin<=end):
    sentence = str(table.cell(begin, 3).value)
    list.append(sentence)
    begin = begin + 1

begin = 2
end = 22349
doc_vec = bc.encode(list)

vlist = {}
for i in doc_vec:
    #table.cell(begin, 7).value = str(i)
    print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")
    print(i[0])
    variablex = torch.tensor(i[0], dtype=torch.double)
    key = str(table.cell(begin, 1).value)
    vlist[key] = variablex
    begin = begin + 1

data.save("bertsentence.xlsx")

with open("douyin_vbert.pkl", 'wb') as fo:  # 将数据写入pkl文件
    pickle.dump(vlist, fo)
"""
begin = 2
end = 20
while(begin<=end):
    pt = list(table.cell(begin,6).value)
    temp = np.array(pt)
    print(temp)
    begin = begin + 1"""